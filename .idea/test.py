from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet("Mysheet",0)
ws.cell(row=4, column=2,value=6)
wb.save('test.xlsx')