from math import sin, cos
import matplotlib.pyplot as plt
from openpyxl import Workbook

fig = plt.figure()
ax = fig.add_subplot(111)

k=2.5

x0 = 0 #int(input("Введите левую границу интервала: "))
xN = 3 #int(input("Введите правую границу интервала: "))

def u(x):
    return sin(2*x)-cos(x)

def du(x):
    return 2 * cos(2 * x) + sin(x)

def f(x,y):
    return du(x)+k*(y-u(x))

def PlotU(x0,fig):
    x = [x0]
    y=[u(x0)]
    for i in range(N):
        x0 += h
        x.append(x0)
        y.append(u(x0))
    au = fig.add_subplot(111)
    au.plot(x, y, 'o-')
   ## plt.show()

def Output(x0,N,h):
    f = open("euler1.txt", "w")
    f.write("Function: "+'\n')
    for i in range(N):
        f.write(str(u(x0)) + " ")
        x0 +=h
    f.close()

def Euler(h,x0,N):

    y0=u(x0)
    y=[y0]
    x=x0
    for i in range (N):
        yi = y0 + h * f(x, y0)
        y.append(yi)
        x=x+h
        y0=yi

    return y

def EulerOut(y):
    f = open("euler.txt", "a")
    f.write('\n'+"Euler: "+'\n')
    for item in y:
        f.write(str(item) + " ")
    f.close()

def EulerPlot(x0,y,fig):
    x=[x0]
    for i in range (N):
        x0+=h
        x.append(x0)
    ax = fig.add_subplot(111)
    ax.plot(x, y, 'o-')
    plt.show()

def CountError(y,x0,h):
    E=[]
    for item in y:
        e=abs(item-u(x0))
        E.append(e)
        x0 +=h
    return max(E)

def ErrorOut(E):
    fl = open("Euler.txt", "a")
    fl.write('\n' + "Error:"+'\n')
    for item in E:
        fl.write(str(item) + " ")
    fl.write('\n'+"Absolute:"+'\n')
    fl.write(str(max(E)))
    fl.close()

def IntoExcel(i,J, h,Em):
    wb = Workbook()
    ws = wb.create_sheet("Mysheet", 0)
    ws.cell(row=4, column=2, value=6)
    wb.save('Euler.xlsx')

def analisys(x0,xN):
    wb = Workbook()
    ws = wb.create_sheet("Mysheet", 0)
    N=5
    s=1
    while N<10000:
        N -=1
        h=(xN-x0)*0.1/N
        ws.cell(row=s, column=1, value=N+1)
        ws.cell(row=s,column=2,value=h)
        ws.cell(row=s, column=3, value=CountError(Euler(h,x0,N),x0,h))
        N=(N+1)*2
        s+=1
    wb.save('euler.xlsx')
    return 0


##PlotU(x0,fig)

N=10
h=(xN-x0)*1.0/N

Output(x0,N,h)



#analisys(x0,xN)

print( (Euler(h,x0,N)))
y=Euler(h,x0,N)
#EulerOut(y)
##ErrorOut(CountError(y,x0,h))
EulerPlot(x0,y,fig)




##print (PK(h,x0,N))













